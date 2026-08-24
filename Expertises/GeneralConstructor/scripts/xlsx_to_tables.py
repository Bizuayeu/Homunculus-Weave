#!/usr/bin/env python
"""めぐる標準単価表 xlsx から v2 単価テーブル JSON を生成する（ローカル専用ツール）

usage: python scripts/xlsx_to_tables.py <xlsx> <out_dir>

`<out_dir>` は GeneralConstructor ルート。生成物は所定の相対パスへ書く：

    python/data/建築単価帯域.json          12 帯の基準値（共同住宅・2.5〜8m・べた）
    python/data/単価オフセット.json        道路 3 / 基礎形状 2 / 種別 2
    tests/fixtures/unit_price_golden.json  `参照用` 78 行のフラット写し

`参照用` 78 行は「帯域基準値 + 道路 + 基礎形状 + 種別」に分解できる（2026-08-24 検証）。
分解が破れる行が 1 つでもあれば、その行を stderr に出して非ゼロ終了する
（岡田氏の将来更新で規則が崩れたら気付くための番人）。番人を通した後、`BAND_OVERRIDES`
の裁定分を帯域と写しの双方へ載せる（xlsx を直に書き換えず、乖離を一箇所に集める）。
オプション単価・定数は `岡田お試し3` から読まず手起こしで管理する（適用条件が数式に
現れないため。出所は各 JSON の metadata）。

openpyxl は dev extra。CI は xlsx を開かず、コミット済みの生成物を検証する。
"""
from __future__ import annotations

import json
import re
import sys
from pathlib import Path

import openpyxl

SHEET = "参照用"
FIRST_ROW = 4
LAST_ROW = 81
EXPECTED_ROWS = 78
EXPECTED_BANDS = 12

AS_OF = "2026-08-04"
SOURCE = "めぐる標準単価表・修正版.xlsx『参照用』A4:J81"
GENERATED_BY = "scripts/xlsx_to_tables.py"

# 列（1-indexed）
COL_種別, COL_下限, COL_帯ラベル, COL_道路, COL_基礎, COL_単価 = 1, 2, 3, 4, 6, 7

# 表記揺れの正規化表。長屋行は共同行とラベルの書き方が違う（矢印の位置）。
ROAD_LABELS = {
    "前面道路、進入路→2.5m以下": "2.5m以下",
    "前面道路→進入路2.5m以下": "2.5m以下",
    "前面道路、進入路→2.5m～8ｍ": "2.5〜8m",
    "前面道路→進入路2.5m～8ｍ": "2.5〜8m",
    "前面道路→幹線道路、バス通り": "幹線・バス通り",
}
HOUSING_LABELS = {"共同基準": "共同住宅", "長屋": "長屋"}
FOUNDATION_LABELS = {"べた基礎": "べた", "杭基礎": "杭"}

# 分解規則のオフセット（万円/㎡）。基準の組み合わせは 共同住宅 × 2.5〜8m × べた。
ROAD_OFFSET = {"2.5m以下": 2, "2.5〜8m": 0, "幹線・バス通り": 1}
FOUNDATION_OFFSET = {"べた": 0, "杭": 2}
HOUSING_OFFSET = {"共同住宅": 0, "長屋": 2}

# 裁定 override（帯の下限 → 基準値）。xlsx の値を大環主の裁定で上書きする。
# 500〜 帯: xlsx は 51 のまま。岡田氏が xlsx を直すまで乖離は意図的に残す
# （出所: 2026-08-24 大環主裁定）。
BAND_OVERRIDES = {500: 52}
OVERRIDE_NOTE = "2026-08-24 大環主裁定 500〜帯 51→52（xlsx 未反映）"


class ConversionError(Exception):
    """xlsx が想定の形でない（読み替えず落とす）"""


def _normalize(raw: object, table: dict[str, str], kind: str, row: int) -> str:
    key = raw if isinstance(raw, str) else ""
    if key not in table:
        raise ConversionError(f"行 {row}: 未知の{kind}ラベル {raw!r}")
    return table[key]


def _parse_band(label: object, 下限: object, row: int) -> tuple[int, int | None]:
    """帯ラベル（例 '200~220㎡' / '400㎡~500㎡' / '500㎡~'）を 下限・上限 に開く"""
    if not isinstance(label, str) or not isinstance(下限, (int, float)):
        raise ConversionError(f"行 {row}: 施工面積範囲/下限面積が読めない ({label!r}, {下限!r})")
    numbers = [int(n) for n in re.findall(r"\d+", label)]
    if not numbers or numbers[0] != int(下限):
        raise ConversionError(f"行 {row}: 帯ラベル {label!r} と下限面積 {下限!r} が一致しない")
    上限 = numbers[1] if len(numbers) > 1 else None  # '500㎡~' は上限開放
    return int(下限), 上限


def read_rows(xlsx: Path) -> list[dict]:
    """`参照用` の 78 行を正規化して読む"""
    workbook = openpyxl.load_workbook(xlsx, data_only=True)
    if SHEET not in workbook.sheetnames:
        raise ConversionError(f"シート {SHEET!r} が無い（シート: {workbook.sheetnames}）")
    sheet = workbook[SHEET]
    rows = []
    for row in range(FIRST_ROW, LAST_ROW + 1):
        単価 = sheet.cell(row, COL_単価).value
        if not isinstance(単価, (int, float)):
            raise ConversionError(f"行 {row}: ㎡単価が数値でない ({単価!r})")
        下限, 上限 = _parse_band(
            sheet.cell(row, COL_帯ラベル).value, sheet.cell(row, COL_下限).value, row
        )
        rows.append(
            {
                "種別": _normalize(sheet.cell(row, COL_種別).value, HOUSING_LABELS, "種別", row),
                "施工面積帯": {"下限": 下限, "上限": 上限},
                "道路区分": _normalize(sheet.cell(row, COL_道路).value, ROAD_LABELS, "道路", row),
                "基礎形状": _normalize(
                    sheet.cell(row, COL_基礎).value, FOUNDATION_LABELS, "基礎形状", row
                ),
                "ベース㎡単価": 単価,
            }
        )
    workbook.close()
    if len(rows) != EXPECTED_ROWS:
        raise ConversionError(f"{EXPECTED_ROWS} 行を期待したが {len(rows)} 行だった")
    return rows


def extract_bands(rows: list[dict]) -> list[dict]:
    """オフセット 0 の組み合わせ（共同住宅・2.5〜8m・べた）を帯域基準値として抜く"""
    bands = [
        {
            "下限": r["施工面積帯"]["下限"],
            "上限": r["施工面積帯"]["上限"],
            "基準値": r["ベース㎡単価"],
            "単位": "万円/㎡",
        }
        for r in rows
        if r["種別"] == "共同住宅" and r["道路区分"] == "2.5〜8m" and r["基礎形状"] == "べた"
    ]
    bands.sort(key=lambda b: b["下限"])
    下限一覧 = [b["下限"] for b in bands]
    if len(bands) != EXPECTED_BANDS or len(set(下限一覧)) != EXPECTED_BANDS:
        raise ConversionError(f"帯域 {EXPECTED_BANDS} 本を期待したが {下限一覧} だった")
    return bands


def verify(rows: list[dict], bands: list[dict]) -> list[str]:
    """全行で 基準値＋オフセット＝Excel 値 を検算し、破れた行の説明を返す"""
    基準値 = {b["下限"]: b["基準値"] for b in bands}
    破れ = []
    for r in rows:
        下限 = r["施工面積帯"]["下限"]
        if 下限 not in 基準値:
            破れ.append(f"{r}: 帯域基準値に下限 {下限} が無い")
            continue
        expected = (
            基準値[下限]
            + ROAD_OFFSET[r["道路区分"]]
            + FOUNDATION_OFFSET[r["基礎形状"]]
            + HOUSING_OFFSET[r["種別"]]
        )
        if expected != r["ベース㎡単価"]:
            破れ.append(f"{r}: 分解値 {expected} != Excel 値 {r['ベース㎡単価']}")
    return 破れ


def apply_overrides(rows: list[dict], bands: list[dict]) -> None:
    """裁定 override を帯域基準値と『参照用』の写しの両方へ効かせる（破壊的更新）

    verify（分解規則の番人）が生の xlsx 値で走り終えた後に呼ぶ。先に上書きすると
    番人が裁定分の差を「規則の破れ」と読んでしまう。
    """
    for band in bands:
        新基準値 = BAND_OVERRIDES.get(int(band["下限"]))
        if 新基準値 is None:
            continue
        差 = 新基準値 - band["基準値"]
        band["基準値"] = 新基準値
        for row in rows:
            if row["施工面積帯"]["下限"] == band["下限"]:
                row["ベース㎡単価"] += 差
                row["override"] = OVERRIDE_NOTE


def _metadata(description: str, *, override: bool = False) -> dict:
    """生成物の metadata（override＝裁定で xlsx の値を上書きした表だけ真）"""
    meta = {
        "description": description,
        "as_of": AS_OF,
        "source": SOURCE,
        "generated_by": GENERATED_BY,
    }
    if override:
        meta["override"] = OVERRIDE_NOTE
    return meta


def build_offsets() -> dict:
    excel_labels: dict[str, list[str]] = {v: [] for v in ROAD_OFFSET}
    for raw, normalized in ROAD_LABELS.items():
        excel_labels[normalized].append(raw)
    return {
        "単価オフセット": {
            "道路区分": [
                {"区分": 区分, "オフセット": 値, "単位": "万円/㎡", "excel_ラベル": excel_labels[区分]}
                for 区分, 値 in ROAD_OFFSET.items()
            ],
            "基礎形状": [
                {"区分": 区分, "オフセット": 値, "単位": "万円/㎡"}
                for 区分, 値 in FOUNDATION_OFFSET.items()
            ],
            "種別": [
                {"区分": "共同住宅", "オフセット": 0, "単位": "万円/㎡"},
                {
                    "区分": "長屋",
                    "オフセット": 2,
                    "単位": "万円/㎡",
                    "derived": True,
                    "note": "Excel に長屋行があるのは 400〜500㎡ 帯のみ。他帯は共同同帯 +2 で導出（2026-08-24 裁定、岡田氏へ事後確認）",
                },
            ],
        },
        "metadata": _metadata(
            "ベース㎡単価の加算オフセット（基準の組み合わせ＝共同住宅・2.5〜8m・べた）"
        ),
    }


def write_json(path: Path, payload: dict) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with open(path, "w", encoding="utf-8", newline="\n") as f:
        json.dump(payload, f, ensure_ascii=False, indent=2)
        f.write("\n")


def main(argv: list[str]) -> int:
    if len(argv) != 3:
        print(f"usage: python {argv[0]} <xlsx> <out_dir>", file=sys.stderr)
        return 2
    xlsx, out_dir = Path(argv[1]), Path(argv[2])
    if not xlsx.exists():
        print(f"xlsx が見つからない: {xlsx}", file=sys.stderr)
        return 2

    try:
        rows = read_rows(xlsx)
        bands = extract_bands(rows)
    except ConversionError as e:
        print(f"変換エラー: {e}", file=sys.stderr)
        return 1

    破れ = verify(rows, bands)
    if 破れ:
        print(
            f"分解規則が破れた行が {len(破れ)} 件ある（帯域＋道路＋基礎形状＋種別）:",
            file=sys.stderr,
        )
        for line in 破れ:
            print(f"  {line}", file=sys.stderr)
        return 1

    apply_overrides(rows, bands)  # 番人（verify）の後に裁定分を載せる

    帯域 = out_dir / "python" / "data" / "建築単価帯域.json"
    オフセット = out_dir / "python" / "data" / "単価オフセット.json"
    golden = out_dir / "tests" / "fixtures" / "unit_price_golden.json"

    write_json(
        帯域,
        {
            "建築単価帯域": bands,
            "metadata": _metadata(
                "施工面積帯ごとのベース㎡単価の基準値（共同住宅・道路区分 2.5〜8m・べた基礎。上限 null は開放）",
                override=True,
            ),
        },
    )
    write_json(オフセット, build_offsets())
    write_json(
        golden,
        {
            "unit_price_golden": rows,
            "metadata": _metadata(
                "『参照用』78 行（共同 72＋長屋 6）のフラット写し。分解規則の回帰テスト用"
                "（override 印のある行は裁定で xlsx の値を上書きしている）",
                override=True,
            ),
        },
    )
    print(f"検算 {len(rows)}/{EXPECTED_ROWS} 行 OK（帯域 {len(bands)} 本）")
    for path in (帯域, オフセット, golden):
        print(f"生成: {path}")
    return 0


if __name__ == "__main__":
    sys.exit(main(sys.argv))
